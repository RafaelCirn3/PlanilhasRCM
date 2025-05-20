from django.urls import reverse_lazy
from django.shortcuts import render
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from .models import Servico, RegistroServico, Insumo, MovimentacaoInsumo
from .forms import (
    ServicoForm,
    RegistroServicoForm,
    InsumoForm,
    MovimentacaoInsumoForm,
)
from datetime import date
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from io import BytesIO
from django.http import HttpResponse


def home(request):
    context = {
        'total_servicos': Servico.objects.count(),
        'total_insumos': Insumo.objects.count(),
        'servicos_hoje': RegistroServico.objects.filter(data=date.today()).count(),
        'servicos_hoje_concluidos': RegistroServico.objects.filter(data=date.today()).count(),
        'ultimos_registros': RegistroServico.objects.all().order_by('-movimentacaoinsumo')[:5],
        'ultimas_movimentacoes': MovimentacaoInsumo.objects.all().order_by('-insumo')[:5],
        'ultimos_insumos': Insumo.objects.all().order_by('-nome')[:5],
    }
    return render(request, 'home.html', context)

def index(request):
    return render(request, 'base.html')
# ---------------------- SERVIÇO ----------------------
class ServicoListView(LoginRequiredMixin, ListView):
    model = Servico
    template_name = 'servico/list.html'

class ServicoDetailView(LoginRequiredMixin, DetailView):
    model = Servico
    template_name = 'servico/detail.html'

class ServicoCreateView(LoginRequiredMixin, CreateView):
    model = Servico
    form_class = ServicoForm
    template_name = 'servico/form.html'
    success_url = reverse_lazy('servico_list')

class ServicoUpdateView(LoginRequiredMixin, UpdateView):
    model = Servico
    form_class = ServicoForm
    template_name = 'servico/form.html'
    success_url = reverse_lazy('servico_list')

class ServicoDeleteView(LoginRequiredMixin, DeleteView):
    model = Servico
    template_name = 'servico/confirm_delete.html'
    success_url = reverse_lazy('servico_list')


# ---------------------- REGISTRO DE SERVIÇO ----------------------
class RegistroServicoListView(LoginRequiredMixin, ListView):
    model = RegistroServico
    template_name = 'registroservico/list.html'
    context_object_name = 'registros'

class RegistroServicoDetailView(LoginRequiredMixin, DetailView):
    model = RegistroServico
    template_name = 'registroservico/detail.html'

class RegistroServicoCreateView(LoginRequiredMixin, CreateView):
    model = RegistroServico
    form_class = RegistroServicoForm
    template_name = 'registroservico/form.html'
    success_url = reverse_lazy('registroservico_list')

    def form_valid(self, form):
        form.instance.executado_por = self.request.user
        return super().form_valid(form)

class RegistroServicoUpdateView(LoginRequiredMixin, UpdateView):
    model = RegistroServico
    form_class = RegistroServicoForm
    template_name = 'registroservico/form.html'
    success_url = reverse_lazy('registroservico_list')

class RegistroServicoDeleteView(LoginRequiredMixin, DeleteView):
    model = RegistroServico
    template_name = 'registroservico/confirm_delete.html'
    success_url = reverse_lazy('registroservico_list')


# ---------------------- INSUMO ----------------------
class InsumoListView(LoginRequiredMixin, ListView):
    model = Insumo
    template_name = 'insumo/list.html'

class InsumoDetailView(LoginRequiredMixin, DetailView):
    model = Insumo
    template_name = 'insumo/detail.html'

class InsumoCreateView(LoginRequiredMixin, CreateView):
    model = Insumo
    form_class = InsumoForm
    template_name = 'insumo/form.html'
    success_url = reverse_lazy('insumo_list')

class InsumoUpdateView(LoginRequiredMixin, UpdateView):
    model = Insumo
    form_class = InsumoForm
    template_name = 'insumo/form.html'
    success_url = reverse_lazy('insumo_list')

class InsumoDeleteView(LoginRequiredMixin, DeleteView):
    model = Insumo
    template_name = 'insumo/confirm_delete.html'
    success_url = reverse_lazy('insumo_list')


# ---------------------- MOVIMENTAÇÃO DE INSUMO ----------------------
class MovimentacaoInsumoListView(LoginRequiredMixin, ListView):
    model = MovimentacaoInsumo
    template_name = 'movimentacaoinsumo/list.html'

class MovimentacaoInsumoDetailView(LoginRequiredMixin, DetailView):
    model = MovimentacaoInsumo
    template_name = 'movimentacaoinsumo/detail.html'

class MovimentacaoInsumoCreateView(LoginRequiredMixin, CreateView):
    model = MovimentacaoInsumo
    form_class = MovimentacaoInsumoForm
    template_name = 'movimentacaoinsumo/form.html'
    success_url = reverse_lazy('movimentacaoinsumo_list')

    def form_valid(self, form):
        form.instance.solicitado_por = self.request.user
        form.save()
        return super().form_valid(form)

class MovimentacaoInsumoUpdateView(LoginRequiredMixin, UpdateView):
    model = MovimentacaoInsumo
    form_class = MovimentacaoInsumoForm
    template_name = 'movimentacaoinsumo/form.html'
    success_url = reverse_lazy('movimentacaoinsumo_list')

class MovimentacaoInsumoDeleteView(LoginRequiredMixin, DeleteView):
    model = MovimentacaoInsumo
    template_name = 'movimentacaoinsumo/confirm_delete.html'
    success_url = reverse_lazy('movimentacaoinsumo_list')

# ---------------------- GERADOR DE EXCEL ----------------------

def excel_registros(request):
    registros = RegistroServico.objects.all().values(
        'id', 
        'servico__nome', 
        'data', 
        'executado_por__username'
    )

    # Criar DataFrame
    df = pd.DataFrame.from_records(registros)
    df.columns = ['ID', 'Serviço', 'Data', 'Executado Por']

    # Formatar coluna Data para string com formato desejado
    df['Data'] = pd.to_datetime(df['Data']).dt.strftime('%d/%m/%Y %H:%M')

    # Criar buffer em memória
    output = BytesIO()

    # Exportar para Excel (sem índice)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Registros de Serviço', index=False)

        # Acessar workbook e worksheet para estilizar
        wb = writer.book
        ws = writer.sheets['Registros de Serviço']

        # Estilos básicos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill("solid", fgColor="4F81BD")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Estilizar cabeçalho
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Ajustar largura das colunas baseado no conteúdo
        for i, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(col)
            )
            ws.column_dimensions[get_column_letter(i)].width = max_length + 5

        # Estilizar dados: centralizar e borda
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.alignment = center_alignment
                cell.border = thin_border

    # Preparar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=registros.xlsx'
    response.write(output.getvalue())
    return response


def excel_movimentacoes(request):
    movimentacoes = MovimentacaoInsumo.objects.all().values(
        'id', 'insumo__nome', 'quantidade', 'finalidade', 'solicitado_por__username'
    )

    df = pd.DataFrame.from_records(movimentacoes)
    df.columns = ['ID', 'Insumo', 'Quantidade', 'Finalidade', 'Solicitado Por']

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Movimentações de Insumo', index=False)

        wb = writer.book
        ws = writer.sheets['Movimentações de Insumo']

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill("solid", fgColor="4F81BD")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        for i, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(col)
            )
            ws.column_dimensions[get_column_letter(i)].width = max_length + 5

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.alignment = center_alignment
                cell.border = thin_border

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=movimentacoes.xlsx'
    response.write(output.getvalue())
    return response